from openpyxl import load_workbook
from itertools import combinations
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np

data_file = 'data/Large_Group_Pricing_Exercise_Solution.xlsx'

#load the entire workbook
wb = load_workbook(data_file, data_only=True)

#Test if loading works by listing all sheets in the file.
# print("Found the following worksheets:")
# for sheetname in wb.sheetnames:
#     print(sheetname)

# Access worksheets and attribute each one an object
ws_gen_info = wb['General Information']
ws_prev_car_info = wb['Previous Carrier Info']
ws_plan_des_info = wb['Plan Design Information']
ws_curr_cens = wb['Current Census']
ws_factor_tables = wb['Factor Tables']
ws_solution = wb['Solution']

gen_info_all_rows = list(ws_gen_info.rows)
solution_all_rows = list(ws_solution.rows)

# add all subscriber ID's to an overall array
sub_id_arr = []
for row in solution_all_rows[9:119]:
    if row[1].value not in sub_id_arr:
        sub_id_arr.append(row[1].value)

# calculate manual rates by excluding subscriber id's a,b,c,d, and e 
# returns 3x4 array of Plans A,B,C and plan designs Single, Single + Spouse, Single + Children, Family
def manual_rate_exclude_five(a,b,c,d,e):
    people_counter = 0
    id_check = [a,b,c,d,e]
    AG_Factor = 0
    Area_Factor = 0
    PB_Factor = 0
    for row in solution_all_rows[9:119]:
        if row[1].value not in id_check:
            AG_Factor += row[10].value
            Area_Factor += row[11].value
            PB_Factor += row[12].value
            people_counter += 1
    manual_rate = gen_info_all_rows[7][2].value * (AG_Factor/people_counter) * (Area_Factor/people_counter) * (PB_Factor/people_counter)
    return manual_rate

#find info for all combos of five subscribers excluded

combination_list = list(combinations(sub_id_arr,5))
c_list_len = len(combination_list)

all_rates_list = []

for i in range(len(combination_list)):
    print(i/c_list_len)
    rate = manual_rate_exclude_five(combination_list[i][0],combination_list[i][1],combination_list[i][2],combination_list[i][3],combination_list[i][4])
    all_rates_list.append(rate)

np.save('all_rates_list.npy', all_rates_list) 

loaded_arr = np.load('all_rates_list.npy')

print(f'min rate {np.min(all_rates_list)}')
print(f'max rate {np.max(all_rates_list)}')
print(f'mean: {np.mean(all_rates_list)}')
print(f'standard dev: {np.std(all_rates_list)}')
print(f'5th percentile {np.percentile(all_rates_list,5)}')
print(f'95th percentile {np.percentile(all_rates_list,95)}')
print(f'2.5th percentile {np.percentile(all_rates_list,2.5)}')
print(f'97.5th percentile {np.percentile(all_rates_list,97.5)}')
sns.displot(all_rates_list, kde=True, bins=20)
